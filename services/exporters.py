"""Specialized export formatters — Single Responsibility Principle."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from services.chart_builder import build_chart_specs
from utils.helpers import ensure_dirs, slugify, utc_now_iso
from utils.logging_config import get_logger

logger = get_logger(__name__)

_BRAND = "0D9488"
_BRAND_DARK = "0F172A"
_BRAND_SOFT = "CCFBF1"
_ROW_ALT = "F8FAFC"
_BORDER = "E2E8F0"
_MUTED = "64748B"
_INK = "0B1220"


def _base(export_dir: Path, label: str, suffix: str) -> Path:
    stamp = utc_now_iso().replace(":", "-").replace(".", "-")[:19]
    return export_dir / f"{slugify(label)}_{stamp}{suffix}"


class CSVExporter:
    """Raw tabular dump — identical values, no styling."""

    def __init__(self, export_dir: str | Path) -> None:
        self.export_dir = Path(export_dir)
        ensure_dirs(self.export_dir)

    def run(self, df: pd.DataFrame, label: str) -> str:
        path = _base(self.export_dir, label, ".csv")
        df.to_csv(path, index=False)
        return str(path)


class JSONExporter:
    def __init__(self, export_dir: str | Path) -> None:
        self.export_dir = Path(export_dir)
        ensure_dirs(self.export_dir)

    def run(self, df: pd.DataFrame, label: str) -> str:
        path = _base(self.export_dir, label, ".json")
        df.to_json(path, orient="records", indent=2, date_format="iso")
        return str(path)


class MarkdownExporter:
    def __init__(self, export_dir: str | Path) -> None:
        self.export_dir = Path(export_dir)
        ensure_dirs(self.export_dir)

    def run(self, df: pd.DataFrame, label: str) -> str:
        path = _base(self.export_dir, label, ".md")
        try:
            md = df.to_markdown(index=False)
        except Exception:  # noqa: BLE001
            md = df.to_string(index=False)
        path.write_text(f"# {label}\n\n{md}\n", encoding="utf-8")
        return str(path)


class ExcelReportExporter:
    """Excel report builder: Executive Summary · raw SQL Result · Analytics Dashboard."""

    def __init__(self, export_dir: str | Path) -> None:
        self.export_dir = Path(export_dir)
        ensure_dirs(self.export_dir)

    def run(
        self,
        df: pd.DataFrame,
        label: str,
        *,
        question: str = "",
        sql: str = "",
        insights: str = "",
        recommendations: str = "",
        meta: dict[str, Any] | None = None,
    ) -> str:
        from openpyxl import Workbook
        from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows

        meta = meta or {}
        path = _base(self.export_dir, label or "sqlmind_report", ".xlsx")
        wb = Workbook()
        q_text = (question or label or "").strip() or "Analytics query"
        if q_text.lower() in {"sqlmind report", "report", "analytics report", "export"}:
            q_text = (question or meta.get("question") or "Analytics query").strip()

        thin = Border(
            left=Side(style="thin", color=_BORDER),
            right=Side(style="thin", color=_BORDER),
            top=Side(style="thin", color=_BORDER),
            bottom=Side(style="thin", color=_BORDER),
        )
        header_fill = PatternFill("solid", fgColor=_BRAND)
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        title_font = Font(name="Calibri", size=22, bold=True, color=_BRAND_DARK)
        brand_font = Font(name="Calibri", size=14, bold=True, color=_BRAND)
        label_font = Font(name="Calibri", size=9, bold=True, color=_MUTED)
        body_font = Font(name="Calibri", size=11, color=_BRAND_DARK)
        soft_fill = PatternFill("solid", fgColor=_BRAND_SOFT)
        banner_fill = PatternFill("solid", fgColor=_INK)
        logo_fill = PatternFill("solid", fgColor=_BRAND)

        # ── Sheet 1: Executive Summary ──
        ws = wb.active
        ws.title = "Executive Summary"
        ws.sheet_view.showGridLines = False
        for col in range(1, 8):
            ws.cell(row=1, column=col).fill = banner_fill
        ws.merge_cells("A1:G1")
        ws.row_dimensions[1].height = 8
        ws.merge_cells("A3:B5")
        ws["A3"] = "SQ"
        ws["A3"].fill = logo_fill
        ws["A3"].font = Font(name="Calibri", size=28, bold=True, color="FFFFFF")
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
        for r in range(3, 6):
            for c in range(1, 3):
                ws.cell(row=r, column=c).fill = logo_fill
                ws.cell(row=r, column=c).border = thin
        ws["C3"] = "SQLMind AI"
        ws["C3"].font = brand_font
        ws["C4"] = "Business Analytics Report"
        ws["C4"].font = title_font
        ws["C5"] = f"Generated {utc_now_iso()[:19].replace('T', ' ')} UTC"
        ws["C5"].font = Font(name="Calibri", size=9, color=_MUTED)
        ws["A7"] = "QUESTION"
        ws["A7"].font = label_font
        ws["A8"] = q_text
        ws["A8"].font = Font(name="Calibri", size=13, bold=True, color=_BRAND_DARK)
        ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A8:G9")
        facts = [
            ("DATABASE", meta.get("database", "—")),
            ("EXECUTION TIME", meta.get("time_s", "—")),
            ("ROWS", meta.get("rows", len(df) if df is not None else 0)),
            ("GENERATED", utc_now_iso()[:19].replace("T", " ")),
        ]
        for i, (k, v) in enumerate(facts):
            ck = ws.cell(row=11, column=1 + i, value=k)
            ck.font = label_font
            ck.fill = soft_fill
            ck.border = thin
            cv = ws.cell(row=12, column=1 + i, value=v if v not in (None, "") else "—")
            cv.font = Font(name="Calibri", size=13, bold=True, color=_BRAND_DARK)
            cv.fill = soft_fill
            cv.border = thin
        insight_text = (insights or "See SQL Result sheet for the dataset.").strip()
        ws["A14"] = "AI SUMMARY"
        ws["A14"].font = label_font
        ws["A15"] = insight_text
        ws["A15"].font = body_font
        ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A15:G19")
        ws["A21"] = "INSIGHTS"
        ws["A21"].font = label_font
        ws["A22"] = insight_text
        ws["A22"].font = body_font
        ws["A22"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A22:G25")
        rec = (recommendations or "").strip() or "Review Analytics Dashboard charts derived from SQL Result values."
        ws["A27"] = "RECOMMENDATIONS"
        ws["A27"].font = label_font
        ws["A28"] = rec
        ws["A28"].font = body_font
        ws["A28"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A28:G31")
        if sql:
            ws["A33"] = "SQL USED"
            ws["A33"].font = label_font
            ws["A34"] = sql.strip()
            ws["A34"].font = Font(name="Consolas", size=9, color=_MUTED)
            ws["A34"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells("A34:G37")
        ws.column_dimensions["A"].width = 22
        for col_idx in range(2, 8):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        # ── Sheet 2: SQL Result — RAW only (identical to CSV) ──
        ws_raw = wb.create_sheet("SQL Result")
        if df is not None and not df.empty:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    # No fills, borders, fonts, merges — plain values only
                    ws_raw.cell(row=r_idx, column=c_idx, value=value)
        else:
            ws_raw.cell(row=1, column=1, value="")

        # ── Sheet 3: Analytics Dashboard — native Excel charts from SQL values ──
        ws_dash = wb.create_sheet("Analytics Dashboard")
        ws_dash.sheet_view.showGridLines = False
        ws_dash["A1"] = "SQLMind AI · Analytics Dashboard"
        ws_dash["A1"].font = title_font
        ws_dash["A2"] = "Charts use exact SQL Result values (not AI summary text)."
        ws_dash["A2"].font = Font(name="Calibri", size=10, color=_MUTED)

        specs = build_chart_specs(df, max_charts=6)
        # Prefer diversity: bar, hbar, pie, line + optional area/column duplicates
        for idx, spec in enumerate(specs):
            start_col = 1 + (idx % 2) * 12
            start_row = 4 + (idx // 2) * 24
            ws_dash.cell(row=start_row, column=start_col, value=spec.title).font = Font(
                name="Calibri", size=12, bold=True, color=_BRAND_DARK
            )
            h1 = ws_dash.cell(row=start_row + 1, column=start_col, value="Label")
            h2 = ws_dash.cell(row=start_row + 1, column=start_col + 1, value="Value")
            h1.font = header_font
            h1.fill = header_fill
            h2.font = header_font
            h2.fill = header_fill
            for r_i, (_, row) in enumerate(spec.data.iterrows()):
                ws_dash.cell(row=start_row + 2 + r_i, column=start_col, value=str(row["label"]))
                ws_dash.cell(row=start_row + 2 + r_i, column=start_col + 1, value=float(row["value"]))
            n = len(spec.data)
            data_ref = Reference(
                ws_dash, min_col=start_col + 1, min_row=start_row + 1, max_row=start_row + 1 + n
            )
            cats = Reference(
                ws_dash, min_col=start_col, min_row=start_row + 2, max_row=start_row + 1 + n
            )
            anchor = f"{get_column_letter(start_col + 3)}{start_row}"
            try:
                if spec.kind == "pie":
                    chart = PieChart()
                    chart.title = spec.title
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats)
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showPercent = True
                    chart.dataLabels.showCatName = True
                    chart.width = 12
                    chart.height = 8
                elif spec.kind == "line":
                    chart = LineChart()
                    chart.title = spec.title
                    chart.style = 10
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats)
                    chart.width = 14
                    chart.height = 8
                elif spec.kind == "hbar":
                    chart = BarChart()
                    chart.type = "bar"
                    chart.style = 10
                    chart.title = spec.title
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats)
                    chart.width = 14
                    chart.height = 8
                else:
                    # column / bar
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = spec.title
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats)
                    chart.width = 14
                    chart.height = 8
                ws_dash.add_chart(chart, anchor)
            except Exception as exc:  # noqa: BLE001
                logger.info("Excel dashboard chart skipped (%s): %s", spec.kind, exc)

        # Extra area chart when we have a line-capable series
        line_specs = [s for s in specs if s.kind == "line"]
        if line_specs:
            spec = line_specs[0]
            start_col = 1
            start_row = 4 + ((len(specs) + 1) // 2) * 24
            ws_dash.cell(row=start_row, column=start_col, value=f"Area · {spec.title}").font = Font(
                name="Calibri", size=12, bold=True, color=_BRAND_DARK
            )
            ws_dash.cell(row=start_row + 1, column=start_col, value="Label").fill = header_fill
            ws_dash.cell(row=start_row + 1, column=start_col, value="Label").font = header_font
            ws_dash.cell(row=start_row + 1, column=start_col + 1, value="Value").fill = header_fill
            ws_dash.cell(row=start_row + 1, column=start_col + 1, value="Value").font = header_font
            for r_i, (_, row) in enumerate(spec.data.iterrows()):
                ws_dash.cell(row=start_row + 2 + r_i, column=start_col, value=str(row["label"]))
                ws_dash.cell(row=start_row + 2 + r_i, column=start_col + 1, value=float(row["value"]))
            n = len(spec.data)
            data_ref = Reference(
                ws_dash, min_col=start_col + 1, min_row=start_row + 1, max_row=start_row + 1 + n
            )
            cats = Reference(
                ws_dash, min_col=start_col, min_row=start_row + 2, max_row=start_row + 1 + n
            )
            try:
                area = AreaChart()
                area.title = f"Area · {spec.title}"
                area.style = 10
                area.add_data(data_ref, titles_from_data=True)
                area.set_categories(cats)
                area.width = 14
                area.height = 8
                ws_dash.add_chart(area, f"{get_column_letter(start_col + 3)}{start_row}")
            except Exception as exc:  # noqa: BLE001
                logger.info("Excel area chart skipped: %s", exc)

        if not specs:
            ws_dash["A4"] = "No chartable categorical/numeric columns in this SQL result."

        wb.save(path)
        return str(path)


class PDFReportExporter:
    """PDF report builder — enterprise report with vector charts from SQL values."""

    def __init__(self, export_dir: str | Path) -> None:
        self.export_dir = Path(export_dir)
        ensure_dirs(self.export_dir)

    def run(
        self,
        df: pd.DataFrame,
        label: str,
        *,
        question: str = "",
        sql: str = "",
        insights: str = "",
        recommendations: str = "",
        meta: dict[str, Any] | None = None,
    ) -> str:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            HRFlowable,
            KeepTogether,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        from services.chart_builder import draw_reportlab_chart, build_chart_specs

        meta = meta or {}
        path = _base(self.export_dir, label or "report", ".pdf")
        doc = SimpleDocTemplate(
            str(path),
            pagesize=letter,
            leftMargin=0.65 * inch,
            rightMargin=0.65 * inch,
            topMargin=0.6 * inch,
            bottomMargin=0.6 * inch,
        )
        styles = getSampleStyleSheet()
        brand = colors.HexColor(f"#{_BRAND}")
        ink = colors.HexColor(f"#{_BRAND_DARK}")
        muted = colors.HexColor(f"#{_MUTED}")
        cover_title = ParagraphStyle(
            "CoverTitle", parent=styles["Title"], fontName="Helvetica-Bold",
            fontSize=22, textColor=ink, alignment=TA_LEFT, spaceAfter=6,
        )
        h1 = ParagraphStyle(
            "H1", parent=styles["Heading1"], fontName="Helvetica-Bold",
            fontSize=14, textColor=ink, spaceBefore=10, spaceAfter=6,
        )
        h2 = ParagraphStyle(
            "H2", parent=styles["Heading2"], fontName="Helvetica-Bold",
            fontSize=11, textColor=brand, spaceBefore=8, spaceAfter=4,
        )
        body = ParagraphStyle(
            "Body", parent=styles["Normal"], fontSize=10,
            textColor=colors.HexColor("#334155"), leading=14, alignment=TA_LEFT,
        )
        small = ParagraphStyle("Small", parent=body, fontSize=8, textColor=muted)
        mono = ParagraphStyle(
            "Mono", parent=styles["Code"], fontName="Courier", fontSize=8,
            textColor=colors.HexColor("#1E293B"), leading=11,
            backColor=colors.HexColor("#F1F5F9"),
        )

        def esc(t: str) -> str:
            return (
                str(t or "").replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace("\n", "<br/>")
            )

        generated = utc_now_iso()[:19].replace("T", " ")
        story: list[Any] = []
        header = Table(
            [["SQLMind AI", "Business Analytics Report"]],
            colWidths=[2.4 * inch, 4.6 * inch],
        )
        header.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, 0), brand),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor(f"#{_INK}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("TOPPADDING", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("LEFTPADDING", (0, 0), (-1, 0), 12),
                ]
            )
        )
        story += [
            header,
            Spacer(1, 16),
            Paragraph(label or "Analytics Report", cover_title),
            Paragraph(f"Generated {generated} UTC", small),
            Spacer(1, 8),
            HRFlowable(width="100%", thickness=1.5, color=brand, spaceAfter=12),
            Paragraph("Question", h2),
            Paragraph(esc(question), body),
            Spacer(1, 8),
        ]
        meta_table = Table(
            [
                ["Database", str(meta.get("database", "—"))],
                ["Execution time", str(meta.get("time_s", "—"))],
                ["Rows", str(meta.get("rows", len(df) if df is not None else 0))],
                ["Generated", generated],
            ],
            colWidths=[1.6 * inch, 5.2 * inch],
        )
        meta_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_BRAND_SOFT}")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("TEXTCOLOR", (0, 0), (-1, -1), ink),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{_BORDER}")),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(meta_table)
        story.append(Spacer(1, 14))
        story.append(Paragraph("Executive Summary", h1))
        story.append(Paragraph(esc(insights or "See data preview."), body))
        story.append(Paragraph("Business Insights", h2))
        story.append(Paragraph(esc(insights or "No additional insights."), body))
        story.append(Paragraph("Recommendations", h2))
        story.append(
            Paragraph(
                esc(recommendations or "Review charts below for patterns in the SQL result set."),
                body,
            )
        )
        if sql:
            story.append(Paragraph("SQL Used", h2))
            story.append(Paragraph(esc(sql), mono))
            story.append(Spacer(1, 10))

        specs = build_chart_specs(df, max_charts=5)
        if specs:
            story.append(Paragraph("Charts", h1))
            story.append(Paragraph("Vector charts generated from SQL result values.", small))
            story.append(Spacer(1, 6))
            for spec in specs:
                try:
                    story.append(KeepTogether([draw_reportlab_chart(spec), Spacer(1, 10)]))
                except Exception as exc:  # noqa: BLE001
                    logger.info("PDF chart skipped (%s): %s", spec.kind, exc)

        if df is not None and not df.empty:
            story.append(Paragraph("SQL Result Table", h1))
            preview = df.head(30)
            data = [list(map(str, preview.columns))] + preview.astype(str).values.tolist()
            col_count = len(preview.columns)
            col_w = min(6.8 * inch / max(col_count, 1), 1.6 * inch)
            table = Table(data, repeatRows=1, colWidths=[col_w] * col_count)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), brand),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{_BORDER}")),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{_ROW_ALT}")]),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            story.append(table)

        doc.build(story)
        return str(path)


class ExportManager:
    """Orchestrates specialized exporters — one responsibility each."""

    def __init__(self, export_dir: str | Path) -> None:
        self.export_dir = Path(export_dir)
        ensure_dirs(self.export_dir)
        self.csv_exporter = CSVExporter(self.export_dir)
        self.json_exporter = JSONExporter(self.export_dir)
        self.markdown_exporter = MarkdownExporter(self.export_dir)
        self.excel_exporter = ExcelReportExporter(self.export_dir)
        self.pdf_exporter = PDFReportExporter(self.export_dir)

    def run(
        self,
        df: pd.DataFrame,
        label: str,
        *,
        question: str = "",
        sql: str = "",
        insights: str = "",
        recommendations: str = "",
        meta: dict[str, Any] | None = None,
        include_pdf: bool = True,
    ) -> dict[str, str]:
        paths: dict[str, str] = {}
        meta = meta or {"rows": len(df) if df is not None else 0}
        try:
            paths["csv"] = self.csv_exporter.run(df, label)
        except Exception as exc:  # noqa: BLE001
            logger.warning("CSVExporter failed: %s", exc)
        try:
            paths["excel"] = self.excel_exporter.run(
                df, label, question=question or label, sql=sql,
                insights=insights, recommendations=recommendations, meta=meta,
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("ExcelReportExporter failed: %s", exc)
        try:
            paths["json"] = self.json_exporter.run(df, label)
        except Exception as exc:  # noqa: BLE001
            logger.warning("JSONExporter failed: %s", exc)
        try:
            paths["markdown"] = self.markdown_exporter.run(df, label)
        except Exception as exc:  # noqa: BLE001
            logger.warning("MarkdownExporter failed: %s", exc)
        if include_pdf:
            try:
                paths["pdf"] = self.pdf_exporter.run(
                    df, label, question=question or label, sql=sql,
                    insights=insights, recommendations=recommendations, meta=meta,
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning("PDFReportExporter failed: %s", exc)
        return paths
